###############################################################################################################
#    pyBarcoAudit   Copyright (C) <2020-22>  <Kevin Scott>                                                    #                                                                                                             #    The program will scan a Excel spreadsheet of calibration dates and produce a usage report.               #
#                                                                                                             #
# usage: pyBarcoAudit.py [-h] [-s SOURCE]                                                                     #
#                                                                                                             #
#     For changes see history.txt                                                                             #
#                                                                                                             #
###############################################################################################################
#    Copyright (C) <2020-22>  <Kevin Scott>                                                                   #
#                                                                                                             #
#    This program is free software: you can redistribute it and/or modify it under the terms of the           #
#    GNU General Public License as published by the Free Software Foundation, either Version 3 of the         #
#    License, or (at your option) any later Version.                                                          #
#                                                                                                             #
#    This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY; without        #
#    even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the               #
#    GNU General Public License for more details.                                                             #
#                                                                                                             #
#    You should have received a copy of the GNU General Public License along with this program.               #
#    If not, see <http://www.gnu.org/licenses/>.                                                              #
#                                                                                                             #
###############################################################################################################

import textwrap
import argparse
import colorama
from tqdm import tqdm
from pathlib import Path
from openpyxl import load_workbook

import src.myTimer  as myTimer
import src.myConfig as myConfig
import src.myLogger as myLogger
from src.myLicense import printLongLicense, printShortLicense
from src.dataClasses import Monitor, MonitorResults, ModelResults, SiteResults
from src.dataMapping import *

import src.utils.monitor_utils as monitor
import src.utils.model_utils   as mu
import src.utils.site_utils    as su



############################################################################################ scanWorkBook ######
def scanWorkBook(source, monitors):
    """  Loads the spreadsheet, already been validated.
         Populates the directory monitors with the contents of the spreadsheet.
    """
    #  Load the spreadsheet in read only [don't want to accidentally amend]
    #  and only load data and not formulas.
    workBook = load_workbook(filename=source, read_only=True, data_only=True)

    #  Grab the active worksheet
    workSheet = workBook.active
    logger.info(f"Scanning {workSheet.title} in {source}")

    for row in tqdm(workSheet.iter_rows(min_row = MIN_ROW,
                                        max_row = MAX_ROW,
                                        min_col = MIN_COL,
                                        max_col = MAX_COL,
                                        values_only=True), unit=" rows", ncols=myConfig.NCOLS):

        monitor = Monitor(monitorType    = row[MONITOR_TYPE],
                          serialNumber   = row[SERIAL_NUMBER],
                          CliEng_ID      = row[CE_ID],
                          PCAssetNumber  = row[PC_ASSET_NO],
                          Site           = row[SITE],
                          Location       = row[LOCATION],
                          DisplayTime    = row[DISPLAY_TIME],
                          BacklightTime  = row[BACKLIGHT_TIME],
                          InstalledDate  = row[INSTALLED_DATE],
                          AcceptanceDate = row[ACCEPTANCE_DATE],
                          FirstPPMDate   = row[FIRST_PPM_DATE],
                          SecondPPMDate  = row[SECOND_PPM_DATE],
                          ThirdPPMDate   = row[THIRD_PPM_DATE],
                          FourthPPMDate  = row[FOURTH_PPM_DATE],
                          FifthPPMDate   = row[FIFTH_PPM_DATE],
                          SixthPPMDate   = row[SIXTH_PPM_DATE],
                          PPMDueDate     = row[PPM_DUE_DATE],
                          Status         = row[STATUS],
                          CommentOne     = row[COMMENT_ONE],
                          CommentTwo     = row[COMMENT_TWO])

        monitors.append(monitor)



############################################################################################## parseArgs ######
def main(source, logger):
    """  Main function, calls each separate stage.
            scanWorkBook(source)                        -   scans the spreadsheet and populates the directory monitors.
            parseMonitors(monitors, monitorResults)     -   collates the results.
            printResults(monitorResults)                -   prints the results.
    """
    monitorResults = MonitorResults()     #  Used to hold the results.
    modelResults   = ModelResults()       #  Used to hold the breakdown by model
    siteResults    = SiteResults()        #  Used to hold the breakdown by site
    monitors       = []                   #  Used to hold the monitor data.

    scanWorkBook(source, monitors)

    monitor.parseMonitors(monitors, monitorResults, logger)
    monitor.printMonitorResults(monitorResults)

    mu.parseModel(monitors, modelResults)
    mu.printModelResults(modelResults)

    su.parseSite(monitors, siteResults)
    su.printSiteResults(siteResults)

############################################################################################## parseArgs ######
def parseArgs():
    """  Process the command line arguments.

         Checks the arguments and will exit if not valid.

         Exit Code 0 - program has exited normally, after print version, licence or help.
         Exit Code 1 - No source supplied.
         Exit Code 2 - Source does not exist.
    """
    parser = argparse.ArgumentParser(
        formatter_class = argparse.RawTextHelpFormatter,
        description=textwrap.dedent("""\
        The program will scan a Excel spreadsheet of calibration dates and produce a usage report.
        -----------------------
        """),
        epilog = f" Kevin Scott (C) 2020-22 :: {myConfig.NAME} {myConfig.VERSION}")

    parser.add_argument("-s",  "--source",  type=Path, action="store", default=False, help="Source file.")
    parser.add_argument("-l",  "--license", action="store_true" , help="Print the Software License.")
    parser.add_argument("-v",  "--version", action="store_true" , help="Print the version of the application.")

    args = parser.parse_args()


    if args.version:
        printShortLicense(myConfig.NAME, myConfig.VERSION)
        logger.info(f"End of {myConfig.NAME} V{myConfig.VERSION}: version")
        print("Goodbye.")
        exit(0)

    if args.license:
        printLongLicense(myConfig.NAME, myConfig.VERSION)
        logger.info(f"End of {myConfig.NAME} V{myConfig.VERSION} : Printed Licence")
        print("Goodbye.")
        exit(0)

    if not args.source:
        logger.error("No Source Supplied.")
        print(f"{colorama.Fore.RED}No Source Supplied. {colorama.Fore.RESET}")
        parser.print_help()
        print("Goodbye.")
        exit(1)

    if not args.source.exists():
        logger.error("Source does not exist.")
        print(f"{colorama.Fore.RED}Source does not exist. {colorama.Fore.RESET}")
        parser.print_help()
        print("Goodbye.")
        exit(2)

    return (args.source)

############################################################################################### __main__ ######

if __name__ == "__main__":

    myConfig    = myConfig.Config()                                                # Need to do this first.
    logger      = myLogger.get_logger(f"logs\\{myConfig.NAME}.log")                      # Create the logger.
    timer       = myTimer.Timer()

    timer.Start

    logger.info("-"*100)
    logger.info(f"Start of {myConfig.NAME} {myConfig.VERSION}")

    source = parseArgs()

    main(source, logger)

    timeStop = timer.Stop

    print()
    print(f"{colorama.Fore.CYAN}Completed :: {timeStop} {colorama.Fore.RESET}")
    logger.info(f"Completed :: {timeStop}")
    logger.info(f"End of {myConfig.NAME} {myConfig.VERSION}")
    logger.info("-"*100)

    exit(0)
